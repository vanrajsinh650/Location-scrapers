from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def get_headers(category):
    return ["#", f"{category.title()} Name", "Mobile Number", "Location"]


def style_sheet(ws, headers, data):
    # header style - dark teal
    header_fill = PatternFill(start_color="006666", end_color="006666", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # alternating row fills
    row_fill_1 = PatternFill(start_color="008080", end_color="008080", fill_type="solid")
    row_fill_2 = PatternFill(start_color="007070", end_color="007070", fill_type="solid")
    row_font = Font(name="Calibri", color="FFFFFF", size=10)
    row_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style="thin", color="005555"),
        right=Side(style="thin", color="005555"),
        top=Side(style="thin", color="005555"),
        bottom=Side(style="thin", color="005555"),
    )

    # write headers
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

    # write data rows
    for row_idx, item in enumerate(data, 1):
        fill = row_fill_1 if row_idx % 2 == 1 else row_fill_2
        row_num = row_idx + 1

        values = [
            row_idx,
            item.get("name", ""),
            item.get("phone", "") or "No Info",
            item.get("area", ""),
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.fill = fill
            cell.font = row_font
            cell.border = thin_border
            if col_idx == 1:
                cell.alignment = center_align
            else:
                cell.alignment = row_align

    # column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 30

    # freeze header
    ws.freeze_panes = "A2"

    # row height
    ws.row_dimensions[1].height = 25
    for r in range(2, len(data) + 2):
        ws.row_dimensions[r].height = 22


def save_to_excel(results, filepath, category=""):
    # only include entries that have a phone number
    results = [r for r in results if r.get("phone")]
    if not results:
        return 0

    wb = Workbook()
    headers = get_headers(category) if category else ["#", "Name", "Mobile Number", "Location"]

    # combined sheet
    ws = wb.active
    ws.title = "All Results"
    style_sheet(ws, headers, results)

    # per-area sheets
    by_area = {}
    for r in results:
        by_area.setdefault(r.get("area", "Other"), []).append(r)

    for area, data in by_area.items():
        name = area[:31].replace("/", "-").replace("\\", "-")
        ws2 = wb.create_sheet(title=name)
        style_sheet(ws2, headers, data)

    wb.save(filepath)
